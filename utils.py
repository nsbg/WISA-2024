import os
import pandas as pd

from openpyxl.styles import PatternFill, Font

def get_version():
    version_file = 'version.txt'

    if not os.path.exists(version_file):
        with open(version_file, 'w') as vf:
            vf.write("1")

        return 1

    with open(version_file, 'r') as vf:
        version = int(vf.read().strip())

    new_version = version+1

    with open(version_file, 'w') as vf:
        vf.write(str(new_version))

    return new_version

def apply_excel_style(save_path, target_df):
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            target_df.to_excel(writer, index=False, sheet_name="Results")

            worksheet = writer.sheets["Results"]

            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)

            content_fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            category_map = {
                "H": "hate",
                "HR": "harassment",
                "V": "violence",
                "S": "sexual",
                "SH": "self-harm"
            }

            # Apply styles to cells based on category comparison
            for row in range(2, len(target_df)+2):  # Adjusting for Excel 1-indexing and header row
                human_category = worksheet.cell(row=row, column=target_df.columns.get_loc('human_category')+1).value
                top_1_category = worksheet.cell(row=row, column=target_df.columns.get_loc('top_1_category')+1).value

                # Compare categories and apply fill if they do not match
                if category_map.get(human_category) != top_1_category:
                    worksheet.cell(row=row, column=target_df.columns.get_loc('input_sentence')+1).fill = content_fill