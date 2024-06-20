import os
import sys
import pandas as pd

from tqdm import tqdm
from datetime import datetime

from openpyxl.styles import PatternFill, Font

from classifier import classify_prompts
from search import set_model_name, create_vector_db, find_most_similar

def get_version():
    version_file = 'version.txt'

    if not os.path.exists(version_file):
        with open(version_file, 'w') as vf:
            vf.write("1")

        return 1

    with open(version_file, 'r') as vf:
        version = int(vf.read().strip())

    new_version = version+1

    with open(version_file, 'w') as vf:
        vf.write(str(new_version))

    return new_version

def main(file_name, model_name):
    now = datetime.now()

    save_time = ''.join(str(now.date()).split("-"))
    
    df = pd.read_csv(file_name)

    set_model_name(model_name)

    modelname_for_save = model_name.split("/")[1]

    abnormal_prompts = classify_prompts(df)

    if not abnormal_prompts.empty:
        results_list = []

        index, model, metadata = create_vector_db()

        for prompt in tqdm(abnormal_prompts['prompt']):
            results = find_most_similar(prompt, model, index, metadata)

            combined_result = {
                'input_sentence': prompt, 
                'human_category': abnormal_prompts.loc[abnormal_prompts['prompt'] == prompt, 'category'].values[0]
                }
            
            for result in results:
                combined_result.update(result)

            results_list.append(combined_result)

        save_dir = "./result"

        file_version = get_version()

        file_path = f"{save_dir}/{save_time}_{modelname_for_save}_result_v{file_version}.xlsx"

        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        results_df = pd.DataFrame(results_list)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            results_df.to_excel(writer, index=False, sheet_name="Results")

            worksheet = writer.sheets["Results"]

            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)

            content_fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            category_map = {
                "H": "hate",
                "HR": "harassment",
                "V": "violence",
                "S": "sexual",
                "SH": "self-harm"
            }

            # Apply styles to cells based on category comparison
            for row in range(2, len(results_df) + 2):  # Adjusting for Excel 1-indexing and header row
                human_category = worksheet.cell(row=row, column=results_df.columns.get_loc('human_category')+1).value
                top_1_category = worksheet.cell(row=row, column=results_df.columns.get_loc('top_1_category')+1).value

                # Compare categories and apply fill if they do not match
                if category_map.get(human_category) != top_1_category:
                    worksheet.cell(row=row, column=results_df.columns.get_loc('input_sentence')+1).fill = content_fill
    else:
        print("No abnormal prompts found.")

if __name__ == "__main__":
    file_name = sys.argv[1]
    embed_model_name = sys.argv[2]

    main(file_name, embed_model_name) 